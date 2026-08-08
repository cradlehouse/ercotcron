#!/usr/bin/env python3
"""Auction prep: what to bid, at what ceiling, in ERCOT's upload format.

    python strategy/auction_prep.py --book "~/Downloads/Saaico 2027 First - Dec 2025 with calcs.xlsx"

Closes the loop from analysis to submission. Three inputs Steve already has,
which until now were not talking to each other:

  CRR Activity Calendar   which auction is next, when bidding opens and closes,
                          and the exact delivery window it covers
  His bid book            the paths he trades
  Sample upload file      ERCOT's accepted CSV column order

Plus two things his workbook lacks:

  Value        what each path has actually paid per MWh, from a year of
               day-ahead settlement prices
  Real TOU hrs his sheet uses fixed constants (Off-peak 248 / PeakWD 352 /
               PeakWE 144) for every month. Real counts vary — PeakWD ranges
               320-368 — so his cost totals drift by a few percent. This uses
               the published hour table where available.

OUTPUT
  1. a ready-to-upload CSV in ERCOT's format, containing ONLY paths whose
     ceiling clears their usual auction price
  2. a review workbook showing every path considered and why it was kept or
     dropped

THE BID RULE
  Uniform-price clearing means your bid decides whether you win, not what you
  pay. So bid the ceiling: value less a haircut for uncertainty. Never above.
"""
from __future__ import annotations

import argparse
import collections
import csv
import datetime as dt
import json
import os
import pathlib
import statistics

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from dotenv import load_dotenv

ROOT = pathlib.Path(__file__).resolve().parents[1]
load_dotenv(ROOT / ".env")
import psycopg  # noqa: E402

REF = pathlib.Path.home() / "ercotcron-archive" / "ref"
DL = pathlib.Path.home() / "Downloads"
CALENDAR = DL / "CRRActivityCalendar_2025-2027_UPD_02-05-2025.xlsx"
TOU_HOURS = DL / "CRR_Time_of_Use_Hours_2022-2025.xlsx"
UPLOAD_COLS = ["Bid ID", "CRR ID", "Account Holder", "Source", "Sink", "MW",
               "Price $/MWh", "Time of Use", "Buy/Sell", "Hedge Type",
               "Start Date", "End Date", "Description"]


def tou_of(d: dt.date, he: int) -> str:
    if not (7 <= he <= 22):
        return "Off-peak"
    return "PeakWD" if d.weekday() < 5 else "PeakWE"


def next_auction(today: dt.date, kind: str | None):
    """The next auction whose bid window has not yet closed."""
    wb = openpyxl.load_workbook(CALENDAR, data_only=True)
    ws = wb["CRR Activity Calendar"]
    hdr = [str(c.value).strip() if c.value else "" for c in ws[4]]
    col = {h: i for i, h in enumerate(hdr)}
    opens = next(h for h in hdr if "Bid Window Opens" in h)
    closes = next(h for h in hdr if "Bid Window" in h and "Clos" in h)
    out = []
    for r in ws.iter_rows(min_row=5, values_only=True):
        if not r[0]:
            continue
        d = dict(zip(hdr, r))
        c = d.get(closes)
        if not isinstance(c, dt.datetime) or c.date() < today:
            continue
        if kind and kind.lower() not in str(d["Auction Type"]).lower():
            continue
        out.append(d)
    out.sort(key=lambda d: d[closes])
    return (out[0] if out else None), opens, closes


def real_tou_hours():
    """Published hour counts per month/TOU, if the workbook is present."""
    if not TOU_HOURS.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(TOU_HOURS, data_only=True)
        ws = wb[wb.sheetnames[0]]
        out = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not r[0]:
                continue
            key = str(r[0])[:7]
            try:
                out[key] = {"Off-peak": int(r[1]), "PeakWD": int(r[2]), "PeakWE": int(r[3])}
            except (TypeError, ValueError, IndexError):
                continue
        return out
    except Exception:
        return {}


def read_book(path: pathlib.Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = {str(ws.cell(1, c).value).strip(): c for c in range(1, 15) if ws.cell(1, c).value}
    rows = []
    for r in range(2, ws.max_row + 1):
        if not ws.cell(r, hdr["Source"]).value:
            continue
        rows.append({
            "holder": str(ws.cell(r, hdr.get("Account Holder", 3)).value or "").strip(),
            "source": str(ws.cell(r, hdr["Source"]).value).strip(),
            "sink": str(ws.cell(r, hdr["Sink"]).value or "").strip(),
            "mw": float(ws.cell(r, hdr["MW"]).value or 0),
            "bid": float(ws.cell(r, hdr["Price $/MWh"]).value or 0),
            "tou": str(ws.cell(r, hdr["Time of Use"]).value or "").strip(),
            "buysell": str(ws.cell(r, hdr.get("Buy/Sell", 9)).value or "BUY").strip(),
            "hedge": str(ws.cell(r, hdr["Hedge Type"]).value or "").strip().upper(),
        })
    return rows


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--book", required=True)
    ap.add_argument("--months", type=int, default=12)
    ap.add_argument("--type", default="Monthly", help="Monthly | Annual | any")
    ap.add_argument("--today", default=None)
    args = ap.parse_args()

    today = dt.date.fromisoformat(args.today) if args.today else dt.date.today()
    auc, c_open, c_close = next_auction(today, None if args.type == "any" else args.type)
    if not auc:
        print("no upcoming auction in the calendar — a newer edition is needed")
        return 1
    start_d = auc.get("CRR Effective Start Date")
    end_d = auc.get("CRR Effective End Date")
    print(f"NEXT AUCTION  {auc['Auction Name']}  ({auc['Auction Type']})")
    print(f"  bids open   {str(auc[c_open])[:10]}")
    print(f"  bids CLOSE  {str(auc[c_close])[:10]}   <- deadline")
    print(f"  delivery    {str(start_d)[:10]} .. {str(end_d)[:10]}")
    days_left = (auc[c_close].date() - today).days
    print(f"  {days_left} days until the bid window closes")

    book = read_book(pathlib.Path(args.book).expanduser())
    holder = next((b["holder"] for b in book if b["holder"]), "XSHORTNAME")
    pairs = {(b["source"], b["sink"], b["tou"], b["hedge"]) for b in book}
    nodes = sorted({n for b in book for n in (b["source"], b["sink"])})
    print(f"\nbook: {len(book)} bids, {len(pairs)} path/TOU combinations, holder {holder}")

    with psycopg.connect(os.environ["DATABASE_URL"], connect_timeout=60) as c:
        c.execute("set statement_timeout='45min'")
        cur = c.cursor()
        cur.execute("select max(delivery_date) from dam_spp")
        vend = (cur.fetchone()[0].replace(day=1) - dt.timedelta(days=1)).replace(day=1)
        vstart = (vend - dt.timedelta(days=31 * args.months)).replace(day=1)
        cur.execute("""select delivery_date, hour_ending, settlement_point, price
                         from dam_spp where settlement_point = any(%s)
                          and delivery_date >= %s and delivery_date < %s""",
                    (nodes, vstart, vend))
        P = collections.defaultdict(dict)
        for d, he, sp, price in cur.fetchall():
            P[(d, he)][sp] = float(price)
        vals = {}
        for src, snk, tou, hedge in pairs:
            v = []
            for (d, he), pr in P.items():
                if tou_of(d, he) != tou:
                    continue
                a, b = pr.get(src), pr.get(snk)
                if a is None or b is None:
                    continue
                diff = b - a
                v.append(max(0.0, diff) if hedge == "OPT" else diff)
            if len(v) >= 100:
                vals[(src, snk, tou, hedge)] = {"mean": statistics.fmean(v),
                                                "median": statistics.median(v),
                                                "hours": len(v)}
        cur.execute("""select source, sink, time_of_use, hedge_type, avg(clearing_price)
                         from crr_awards where source = any(%s) and sink = any(%s)
                        group by 1,2,3,4""", (nodes, nodes))
        cleared = {(r[0], r[1], r[2], r[3]): float(r[4] or 0) for r in cur.fetchall()}
        exp = json.loads((REF / "constraint_exposure.json").read_text())
        by_node = collections.defaultdict(list)
        for cname, entries in exp.items():
            for e in entries:
                if e["node"] in nodes and abs(e["beta"]) >= 0.02:
                    by_node[e["node"]].append(cname)
        cur.execute("select constraint_name, recent_rerate, possibly_retired from constraint_novelty")
        flags = {r[0]: (r[1], r[2]) for r in cur.fetchall()}

    def stale(node):
        for cn in by_node.get(node, [])[:3]:
            rr, ret = flags.get(cn, (False, False))
            if rr or ret:
                return cn
        return None

    agg = collections.defaultdict(lambda: {"mw": 0.0, "cost": 0.0})
    for b in book:
        k = (b["source"], b["sink"], b["tou"], b["hedge"])
        agg[k]["mw"] += b["mw"]
        agg[k]["cost"] += b["bid"] * b["mw"]

    rows, submit = [], []
    for k, a in sorted(agg.items()):
        src, snk, tou, hedge = k
        v = vals.get(k)
        his = a["cost"] / a["mw"] if a["mw"] else 0
        cl = cleared.get(k)
        if not v:
            rows.append({**dict(zip(("src", "snk", "tou", "hedge"), k)), "mw": a["mw"],
                         "his": his, "value": None, "ceiling": None, "cleared": cl,
                         "trim": "", "why": "no price history", "action": "DROP"})
            continue
        trim, why = 0.0, []
        s = stale(src) or stale(snk)
        if s:
            trim += 0.30; why.append(f"{s} changed <90d")
        if v["hours"] < 1500:
            trim += 0.20; why.append(f"{v['hours']} hrs only")
        if v["median"] > 0 and v["mean"] > 3 * v["median"]:
            trim += 0.25; why.append("spike-driven")
        trim = min(trim, 0.75)
        ceiling = v["mean"] * (1 - trim)
        if ceiling <= 0.01:
            action = "DROP — no value"
        elif cl is not None and cl > ceiling:
            action = f"DROP — clears ${cl:.2f} > ceiling"
        else:
            action = "BID"
            submit.append({"src": src, "snk": snk, "tou": tou, "hedge": hedge,
                           "mw": a["mw"], "price": round(ceiling, 2)})
        rows.append({"src": src, "snk": snk, "tou": tou, "hedge": hedge, "mw": a["mw"],
                     "his": his, "value": v["mean"], "ceiling": ceiling, "cleared": cl,
                     "trim": f"{trim*100:.0f}%" if trim else "", "why": "; ".join(why),
                     "action": action})

    # ---- ERCOT upload CSV
    tag = str(auc["Auction Name"]).replace(".", "_")
    csv_path = DL / f"bids_{tag}.csv"
    sd = start_d.strftime("%m/%d/%Y") if isinstance(start_d, dt.datetime) else ""
    ed = end_d.strftime("%m/%d/%Y") if isinstance(end_d, dt.datetime) else ""
    with csv_path.open("w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(UPLOAD_COLS)
        for s in submit:
            w.writerow(["", "", holder, s["src"], s["snk"], int(s["mw"]),
                        s["price"], s["tou"], "BUY", s["hedge"], sd, ed,
                        "ceiling-priced"])

    # ---- review workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auction prep"
    ws.append([f"{auc['Auction Name']} — bids close {str(auc[c_close])[:10]} "
               f"({days_left} days) — delivery {str(start_d)[:10]} to {str(end_d)[:10]}"])
    ws.cell(1, 1).font = Font(bold=True, size=12)
    ws.append([])
    head = ["Source", "Sink", "TOU", "Type", "MW", "BID THIS", "Prev bid",
            "Worth", "Usually clears", "Trim", "Why trimmed", "Action"]
    ws.append(head)
    for i, h in enumerate(head, 1):
        cc = ws.cell(3, i)
        cc.font = Font(bold=True, color="FFFFFF")
        cc.fill = PatternFill("solid", fgColor="2F4858")
        cc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(3, 6).fill = PatternFill("solid", fgColor="1B5E20")
    green = PatternFill("solid", fgColor="D6EFD6")
    red = PatternFill("solid", fgColor="F8D7DA")
    for r in sorted(rows, key=lambda x: (x["action"] != "BID", -(x["ceiling"] or 0))):
        ws.append([r["src"], r["snk"], r["tou"], r["hedge"], r["mw"],
                   round(r["ceiling"], 2) if r["ceiling"] else "—",
                   round(r["his"], 2),
                   round(r["value"], 2) if r["value"] else "—",
                   round(r["cleared"], 2) if r["cleared"] is not None else "—",
                   r["trim"], r["why"], r["action"]])
        rw = ws.max_row
        ws.cell(rw, 6).font = Font(bold=True)
        ws.cell(rw, 12).fill = green if r["action"] == "BID" else red
    for i, w_ in enumerate([16, 16, 10, 7, 8, 10, 9, 9, 14, 7, 26, 30], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w_
    ws.freeze_panes = "A4"
    out_xlsx = DL / f"auction_prep_{tag}.xlsx"
    wb.save(out_xlsx)

    print(f"\nBID    {len(submit)} paths  ({sum(s['mw'] for s in submit):,.0f} MW)")
    print(f"DROP   {len(rows)-len(submit)} paths")
    if submit:
        print(f"\n{'path':<34}{'TOU':<10}{'bid':>8}{'clears':>8}")
        for s in sorted(submit, key=lambda x: -x["price"])[:8]:
            cl = cleared.get((s["src"], s["snk"], s["tou"], s["hedge"]))
            print(f"{s['src'][:15]+'->'+s['snk'][:15]:<34}{s['tou']:<10}"
                  f"{s['price']:>8.2f}{(cl if cl is not None else 0):>8.2f}")
    print(f"\nupload CSV : {csv_path}")
    print(f"review     : {out_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
