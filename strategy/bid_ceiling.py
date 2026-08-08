#!/usr/bin/env python3
"""What should I bid? — a ceiling per path, for use before an auction.

    python strategy/bid_ceiling.py --book "~/Downloads/Saaico 2027 First - Dec 2025 with calcs-1.xlsx"

WHY A CEILING IS THE RIGHT OUTPUT

CRR auctions clear at a uniform price. Bid $5, clear at $1, and you pay $1 —
your bid decides whether you win, not what you pay. So there is no cost to
bidding high except one: bidding above what the path is actually worth. That
makes the whole problem "estimate the value, never bid past it."

HOW THE CEILING IS BUILT

  1. VALUE          mean realised congestion per MWh over a trailing window.
                    OPT pays max(0, sink-source); OBL pays the signed difference.

  2. HAIRCUT        the value estimate is not precise, so trim it:
                      -30%  a driving constraint was re-rated in the last 90d
                            (history may describe a network that no longer exists)
                      -20%  thin history (under 1,500 priced hours)
                      -25%  spike-driven (mean more than 3x the median — the
                            average is carried by rare hours, not steady payment)
                    Haircuts stack, floored at 25% of raw value.

  3. CEILING        value after haircut. Bid up to this; never above.

  4. LIKELY COST    what this path has actually cleared at in past auctions.
                    If the ceiling is far above it, the bid is nearly free money
                    and can be placed with confidence.

WHAT THIS DOES NOT DO

It does not predict. It says what a path has been worth and how much of that
estimate deserves trust. Every predictive rule tested on this data failed out
of sample; valuation and risk-flagging survived. Treat the ceiling as a
discipline for not overpaying, not as a forecast of profit.
"""
from __future__ import annotations

import argparse
import collections
import datetime as dt
import json
import os
import pathlib
import statistics

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from dotenv import load_dotenv

ROOT = pathlib.Path(__file__).resolve().parents[1]
load_dotenv(ROOT / ".env")
import psycopg  # noqa: E402

REF = pathlib.Path.home() / "ercotcron-archive" / "ref"
OUT = pathlib.Path.home() / "Downloads" / "steve_bid_ceilings.xlsx"


def tou_of(d: dt.date, he: int) -> str:
    if not (7 <= he <= 22):
        return "Off-peak"
    return "PeakWD" if d.weekday() < 5 else "PeakWE"


def read_book(path: pathlib.Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = {str(ws.cell(1, c).value).strip(): c for c in range(1, 15)
           if ws.cell(1, c).value}
    rows = []
    for r in range(2, ws.max_row + 1):
        src = ws.cell(r, hdr["Source"]).value
        if not src:
            continue
        rows.append({
            "source": str(src).strip(),
            "sink": str(ws.cell(r, hdr["Sink"]).value or "").strip(),
            "mw": float(ws.cell(r, hdr["MW"]).value or 0),
            "bid": float(ws.cell(r, hdr["Price $/MWh"]).value or 0),
            "tou": str(ws.cell(r, hdr["Time of Use"]).value or "").strip(),
            "hedge": str(ws.cell(r, hdr["Hedge Type"]).value or "").strip().upper(),
        })
    return rows


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--book", required=True)
    ap.add_argument("--months", type=int, default=12)
    ap.add_argument("--book-name", default=None)
    args = ap.parse_args()

    book = read_book(pathlib.Path(args.book).expanduser())
    pairs = {(b["source"], b["sink"], b["tou"], b["hedge"]) for b in book}
    nodes = sorted({n for b in book for n in (b["source"], b["sink"])})
    print(f"book: {len(book)} bids across {len(pairs)} path/TOU combinations")

    with psycopg.connect(os.environ["DATABASE_URL"], connect_timeout=60) as c:
        c.execute("set statement_timeout='45min'")
        cur = c.cursor()
        cur.execute("select max(delivery_date) from dam_spp")
        end = (cur.fetchone()[0].replace(day=1) - dt.timedelta(days=1)).replace(day=1)
        start = (end - dt.timedelta(days=31 * args.months)).replace(day=1)
        print(f"valuation window {start} .. {end}")

        # --- value: realised congestion per path/TOU/hedge
        cur.execute("""
            select delivery_date, hour_ending, settlement_point, price
              from dam_spp
             where settlement_point = any(%s)
               and delivery_date >= %s and delivery_date < %s
        """, (nodes, start, end))
        P = collections.defaultdict(dict)
        for d, he, sp, price in cur.fetchall():
            P[(d, he)][sp] = float(price)

        vals = {}
        for src, snk, tou, hedge in pairs:
            v = []
            for (d, he), prices in P.items():
                if tou_of(d, he) != tou:
                    continue
                a, b = prices.get(src), prices.get(snk)
                if a is None or b is None:
                    continue
                diff = b - a
                v.append(max(0.0, diff) if hedge == "OPT" else diff)
            if len(v) >= 100:
                vals[(src, snk, tou, hedge)] = {
                    "mean": statistics.fmean(v),
                    "median": statistics.median(v),
                    "hours": len(v),
                }

        # --- what these paths have actually cleared at
        cur.execute("""
            select source, sink, time_of_use, hedge_type,
                   avg(clearing_price), max(clearing_price), count(distinct auction_name)
              from crr_awards
             where source = any(%s) and sink = any(%s)
             group by 1,2,3,4
        """, (nodes, nodes))
        cleared = {(r[0], r[1], r[2], r[3]): {"avg": float(r[4] or 0),
                                              "max": float(r[5] or 0),
                                              "auctions": r[6]}
                   for r in cur.fetchall()}
        print(f"paths with auction clearing history: {len(cleared)}")

        # --- risk flags from the novelty layer
        exp = json.loads((REF / "constraint_exposure.json").read_text())
        by_node = collections.defaultdict(list)
        for cname, entries in exp.items():
            for e in entries:
                if e["node"] in nodes and abs(e["beta"]) >= 0.02:
                    by_node[e["node"]].append(cname)
        cur.execute("select constraint_name, recent_rerate, possibly_retired from constraint_novelty")
        flags = {r[0]: (r[1], r[2]) for r in cur.fetchall()}

    def stale(node):
        for cname in by_node.get(node, [])[:3]:
            rr, ret = flags.get(cname, (False, False))
            if rr or ret:
                return cname
        return None

    rows = []
    agg = collections.defaultdict(lambda: {"mw": 0.0, "cost": 0.0})
    for b in book:
        k = (b["source"], b["sink"], b["tou"], b["hedge"])
        agg[k]["mw"] += b["mw"]
        agg[k]["cost"] += b["bid"] * b["mw"]

    for k, a in agg.items():
        src, snk, tou, hedge = k
        v = vals.get(k)
        his_bid = a["cost"] / a["mw"] if a["mw"] else 0
        if not v:
            rows.append({"src": src, "snk": snk, "tou": tou, "hedge": hedge,
                         "mw": a["mw"], "his_bid": his_bid, "value": None,
                         "ceiling": None, "cleared": None, "reasons": "no price history",
                         "verdict": "SKIP — cannot value"})
            continue

        haircut, reasons = 0.0, []
        s = stale(src) or stale(snk)
        if s:
            haircut += 0.30
            reasons.append(f"driver {s} changed recently")
        if v["hours"] < 1500:
            haircut += 0.20
            reasons.append(f"only {v['hours']} priced hours")
        if v["median"] > 0 and v["mean"] > 3 * v["median"]:
            haircut += 0.25
            reasons.append("value comes from rare spikes")
        haircut = min(haircut, 0.75)
        ceiling = v["mean"] * (1 - haircut)

        cl = cleared.get(k)
        cl_avg = cl["avg"] if cl else None

        if ceiling <= 0:
            verdict = "SKIP — no value"
        elif cl_avg is not None and cl_avg > ceiling:
            verdict = f"SKIP — clears above ceiling (${cl_avg:.2f})"
        elif his_bid > ceiling:
            verdict = f"LOWER — you bid ${his_bid:.2f}, ceiling ${ceiling:.2f}"
        elif cl_avg is not None and ceiling > cl_avg * 3:
            verdict = "STRONG — wide margin over clearing"
        else:
            verdict = "OK"

        rows.append({"src": src, "snk": snk, "tou": tou, "hedge": hedge,
                     "mw": a["mw"], "his_bid": his_bid, "value": v["mean"],
                     "median": v["median"], "hours": v["hours"],
                     "haircut": haircut, "ceiling": ceiling,
                     "cleared": cl_avg,
                     "headroom": (ceiling / cl_avg) if cl_avg else None,
                     "reasons": "; ".join(reasons) or "full confidence",
                     "verdict": verdict})

    rows.sort(key=lambda r: (r["ceiling"] is None, -(r["ceiling"] or 0)))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bid ceilings"
    head = ["Source", "Sink", "TOU", "Type", "MW",
            "BID UP TO $/MWh", "You bid", "Worth", "Usually clears",
            "Headroom", "Trim %", "Why trimmed", "Verdict"]
    ws.append(head)
    for i, h in enumerate(head, 1):
        cc = ws.cell(1, i)
        cc.font = Font(bold=True, color="FFFFFF")
        cc.fill = PatternFill("solid", fgColor="2F4858")
        cc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(1, 6).fill = PatternFill("solid", fgColor="1B5E20")

    green = PatternFill("solid", fgColor="D6EFD6")
    red = PatternFill("solid", fgColor="F8D7DA")
    amber = PatternFill("solid", fgColor="FFF3CD")
    for r in rows:
        ws.append([
            r["src"], r["snk"], r["tou"], r["hedge"], r["mw"],
            round(r["ceiling"], 2) if r["ceiling"] is not None else "—",
            round(r["his_bid"], 2),
            round(r["value"], 2) if r["value"] is not None else "—",
            round(r["cleared"], 2) if r.get("cleared") is not None else "—",
            f"{r['headroom']:.1f}x" if r.get("headroom") else "—",
            f"{r.get('haircut', 0)*100:.0f}%" if r.get("haircut") else "",
            r["reasons"], r["verdict"],
        ])
        row = ws.max_row
        ws.cell(row, 6).font = Font(bold=True)
        if r["verdict"].startswith("STRONG"):
            ws.cell(row, 13).fill = green
        elif r["verdict"].startswith(("SKIP", "LOWER")):
            ws.cell(row, 13).fill = red
        if r.get("haircut"):
            ws.cell(row, 11).fill = amber
    for i, w in enumerate([16, 16, 10, 7, 8, 15, 9, 9, 14, 10, 8, 34, 34], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    ms = wb.create_sheet("How to use this")
    for line in [
        ["Bid ceilings — how to use this sheet"],
        [""],
        ["THE ONE RULE", "Bid up to the green column. Never above it."],
        [""],
        ["WHY BIDDING HIGH IS SAFE (up to the ceiling)"],
        ["", "CRR auctions clear at a uniform price. If you bid $5 and it clears at $1, you pay $1."],
        ["", "Your bid decides whether you win, not what you pay. So the only mistake is bidding"],
        ["", "above what the path is worth — bidding below the ceiling but above the clearing price"],
        ["", "costs nothing extra and wins you the path."],
        [""],
        ["THE COLUMNS"],
        ["BID UP TO", "Your ceiling. Worth, minus a trim for uncertainty."],
        ["Worth", "What the path has actually paid per MWh over the last 12 months."],
        ["Usually clears", "What this path has cost in past auctions. If the ceiling is well above this,"],
        ["", "you can bid confidently and still win cheaply."],
        ["Headroom", "Ceiling divided by usual clearing price. 3x or more is a wide margin."],
        ["Trim %", "How much was cut from Worth to get the ceiling, and why."],
        [""],
        ["WHY VALUE GETS TRIMMED"],
        ["-30%", "A constraint that drives this path was re-rated in the last 90 days. The history"],
        ["", "may be describing a network that no longer exists."],
        ["-20%", "Thin history — under 1,500 priced hours, so the average is less reliable."],
        ["-25%", "The average is carried by rare spikes rather than steady payment. You may wait"],
        ["", "a long time between payoffs, which is hard on a small book."],
        [""],
        ["WHAT THIS IS NOT"],
        ["", "It is not a prediction. It says what a path has been worth and how much of that"],
        ["", "estimate deserves trust. It stops you overpaying; it cannot promise a profit."],
    ]:
        ms.append(line)
    ms.column_dimensions["A"].width = 16
    ms.column_dimensions["B"].width = 100
    ms.cell(1, 1).font = Font(bold=True, size=14)
    for r in (3, 5, 11, 19, 26):
        ms.cell(r, 1).font = Font(bold=True)
    wb.save(OUT)

    priced = [r for r in rows if r["ceiling"] is not None]
    over = [r for r in priced if r["his_bid"] > r["ceiling"]]
    strong = [r for r in priced if r["verdict"].startswith("STRONG")]
    print(f"\npaths priced: {len(priced)}/{len(rows)}")
    print(f"he bid ABOVE the ceiling on: {len(over)} paths "
          f"({sum(r['mw'] for r in over):,.0f} MW)")
    print(f"wide-margin opportunities: {len(strong)}")
    print(f"\n{'path':<32}{'TOU':<10}{'bid up to':>10}{'he bid':>8}{'clears':>8}")
    for r in priced[:8]:
        print(f"{r['src'][:14]+'->'+r['snk'][:14]:<32}{r['tou']:<10}"
              f"{r['ceiling']:>10.2f}{r['his_bid']:>8.2f}"
              f"{(r['cleared'] if r.get('cleared') is not None else 0):>8.2f}")
    print(f"\nwrote {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
