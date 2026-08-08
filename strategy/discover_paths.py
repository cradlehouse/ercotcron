#!/usr/bin/env python3
"""Discovery: paths worth bidding that are NOT in Steve's book.

    python strategy/discover_paths.py

Candidate generation, not copied conviction. The four firms that beat the CRR
market across nine measured months (Wolframium 9/9, DC Energy 8/9, SESCO and
Peak 6/9) define a candidate universe of paths — positions can be hedges or
offsets, so holding is not evidence of value. Every candidate is then priced by
OUR method: realised congestion over a trailing window, trimmed for
uncertainty, compared against what the auction actually charges for it.

Kept only when:
  ceiling > 1.5x its usual clearing price   (real headroom, not a rounding gap)
  >= 2,000 priced hours                     (not a thin fluke)
  not already in the reference book         (this is the "more than parrot" filter)

Published to path_valuations under book='Discovery', so /bids renders it as its
own section.
"""
from __future__ import annotations

import collections
import datetime as dt
import json
import os
import pathlib
import statistics

import openpyxl
from dotenv import load_dotenv

ROOT = pathlib.Path(__file__).resolve().parents[1]
load_dotenv(ROOT / ".env")
import psycopg  # noqa: E402

REF = pathlib.Path.home() / "ercotcron-archive" / "ref"
BOOK = pathlib.Path.home() / "Downloads" / "Saaico 2027 First - Dec 2025 with calcs.xlsx"
WINNERS = ("XWOLF1", "XWOLF2", "XWOLFP", "XDCEN1", "XDCEN2", "XDCEN3", "XDCEND",
           "XDCENG", "XDCND2", "XSESSW", "XSESCA", "XSESKP", "XPEAK1", "XPEAK2", "XPEAKE")
MAX_CANDIDATES = 120      # keep the price pull bounded — the DB just died once
MIN_HOURS = 2000
HEADROOM = 1.5


def tou_of(d: dt.date, he: int) -> str:
    if not (7 <= he <= 22):
        return "Off-peak"
    return "PeakWD" if d.weekday() < 5 else "PeakWE"


def steve_pairs() -> set[tuple[str, str]]:
    wb = openpyxl.load_workbook(BOOK, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = {str(ws.cell(1, c).value).strip(): c for c in range(1, 15) if ws.cell(1, c).value}
    out = set()
    for r in range(2, ws.max_row + 1):
        s = ws.cell(r, hdr["Source"]).value
        k = ws.cell(r, hdr["Sink"]).value
        if s and k:
            out.add((str(s).strip(), str(k).strip()))
    return out


def main() -> int:
    known = steve_pairs()
    print(f"reference book: {len(known)} source→sink pairs to exclude")

    with psycopg.connect(os.environ["DATABASE_URL"], connect_timeout=60) as c:
        c.execute("set statement_timeout='45min'")
        cur = c.cursor()

        # Candidates: winners' largest OPT positions in monthly auctions, by MW.
        cur.execute("""
            select source, sink, time_of_use, hedge_type,
                   sum(mw) mw, avg(clearing_price) cleared,
                   count(distinct account_holder) holders
              from crr_awards
             where account_holder = any(%s)
               and auction_name like '%%Monthly'
             group by 1,2,3,4
             order by sum(mw) desc
             limit %s
        """, (list(WINNERS), MAX_CANDIDATES * 3))
        cands = []
        for src, snk, tou, hedge, mw, cleared, holders in cur.fetchall():
            if (src, snk) in known:
                continue
            cands.append({"src": src, "snk": snk, "tou": tou, "hedge": hedge,
                          "mw": float(mw), "cleared": float(cleared or 0),
                          "holders": holders})
            if len(cands) >= MAX_CANDIDATES:
                break
        nodes = sorted({n for p in cands for n in (p["src"], p["snk"])})
        print(f"candidates after excluding the book: {len(cands)} paths, {len(nodes)} nodes")

        cur.execute("select max(delivery_date) from dam_spp")
        end = (cur.fetchone()[0].replace(day=1) - dt.timedelta(days=1)).replace(day=1)
        start = (end - dt.timedelta(days=372)).replace(day=1)
        print(f"valuation window {start} .. {end}")

        # ONE bounded price pull for every candidate node.
        cur.execute("""select delivery_date, hour_ending, settlement_point, price
                         from dam_spp where settlement_point = any(%s)
                          and delivery_date >= %s and delivery_date < %s""",
                    (nodes, start, end))
        P = collections.defaultdict(dict)
        for d, he, sp, price in cur.fetchall():
            P[(d, he)][sp] = float(price)
        print(f"price grid: {len(P):,} hours")

        # risk flags for the trim
        exp = json.loads((REF / "constraint_exposure.json").read_text())
        by_node = collections.defaultdict(list)
        for cname, entries in exp.items():
            for e in entries:
                if e["node"] in nodes and abs(e["beta"]) >= 0.02:
                    by_node[e["node"]].append(cname)
        cur.execute("select constraint_name, recent_rerate, possibly_retired from constraint_novelty")
        flags = {r[0]: (r[1], r[2]) for r in cur.fetchall()}

    def stale(node):
        for cn in by_node.get(node, [])[:3]:
            rr, ret = flags.get(cn, (False, False))
            if rr or ret:
                return cn
        return None

    keep = []
    for p in cands:
        vals = []
        for (d, he), pr in P.items():
            if tou_of(d, he) != p["tou"]:
                continue
            a, b = pr.get(p["src"]), pr.get(p["snk"])
            if a is None or b is None:
                continue
            diff = b - a
            vals.append(max(0.0, diff) if p["hedge"] == "OPT" else diff)
        if len(vals) < MIN_HOURS:
            continue
        mean = statistics.fmean(vals)
        med = statistics.median(vals)
        sv = sorted(vals)
        trim, why = 0.0, []
        s = stale(p["src"]) or stale(p["snk"])
        if s:
            trim += 0.30; why.append(f"{s} changed <90d")
        if med > 0 and mean > 3 * med:
            trim += 0.25; why.append("spike-driven")
        trim = min(trim, 0.75)
        ceiling = mean * (1 - trim)
        if p["cleared"] > 0 and ceiling < HEADROOM * p["cleared"]:
            continue
        if ceiling <= 0.05:
            continue
        keep.append({**p, "mean": mean, "median": med,
                     "p05": sv[int(0.05 * len(sv))], "p95": sv[int(0.95 * len(sv))],
                     "pct_pos": 100 * sum(1 for v in vals if v > 0) / len(vals),
                     "hours": len(vals), "trim": trim, "ceiling": ceiling,
                     "why": "; ".join(why)})

    keep.sort(key=lambda r: -(r["ceiling"] - r["cleared"]))
    print(f"\nsurvived valuation + headroom filter: {len(keep)}")
    print(f"{'path':<36}{'TOU':<10}{'ceiling':>8}{'clears':>8}{'held by':>8}")
    for r in keep[:12]:
        print(f"{r['src'][:16]+'->'+r['snk'][:16]:<36}{r['tou']:<10}"
              f"{r['ceiling']:>8.2f}{r['cleared']:>8.2f}{r['holders']:>8}")

    with psycopg.connect(os.environ["DATABASE_URL"], connect_timeout=30) as c:
        with c.cursor() as cur:
            cur.execute("delete from path_valuations where book = 'Discovery'")
            cur.executemany("""
                insert into path_valuations
                  (book, source, sink, time_of_use, hedge_type, mw, bids, bid_price,
                   value_mean, value_median, value_p05, value_p95, pct_hours_pos,
                   hours, edge, drivers, warnings, window_start, window_end,
                   ceiling, cleared_price, trim_pct)
                values ('Discovery',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, [(r["src"], r["snk"], r["tou"], r["hedge"], r["mw"], r["holders"],
                   None, r["mean"], r["median"], r["p05"], r["p95"], r["pct_pos"],
                   r["hours"], r["ceiling"] - r["cleared"],
                   "; ".join(by_node.get(r["snk"], [])[:2]) or None,
                   r["why"] or None, start, end,
                   r["ceiling"], r["cleared"], r["trim"]) for r in keep])
        c.commit()
    print(f"\npublished {len(keep)} discovery paths to path_valuations (book='Discovery')")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
