#!/usr/bin/env python3
"""Backfill unplanned resource outages (NP1-346-ER).

    python scripts/backfill_resource_outages.py --days 1400

Daily unit-level snapshot of forced/maintenance generator outages, published
with a ~3-day disclosure lag. Not fast enough to trade day-of; exactly right
for research: outage state is one of the drivers the constraint framework says
we must condition on, and until now we held none of it.

Transmission outages remain secure-MIS only (Outage Scheduler, CEII). The
public proxy for those is already ingested: RUC binding constraints run on the
outage-adjusted network, so RUC-vs-DAM disagreement carries their shadow.

Each file is a snapshot; rows are keyed by (report_date, unit, start) so the
evolution of one outage across days is visible rather than collapsed.
"""
from __future__ import annotations

import argparse, io, json, os, pathlib, sys, time, zipfile
import urllib.request
from datetime import datetime, timedelta, timezone

import openpyxl
from dotenv import load_dotenv

ROOT = pathlib.Path(__file__).resolve().parents[1]
load_dotenv(ROOT / ".env")
sys.path.insert(0, str(ROOT))

import psycopg  # noqa: E402
from ercot.client import ErcotClient  # noqa: E402

DDL = """
create table if not exists resource_outages (
  report_date    date not null,
  resource_name  text not null,
  unit_code      text not null default '',
  fuel_type      text,
  outage_type    text,
  max_mw         numeric,
  available_mw   numeric,
  reduction_mw   numeric,
  outage_start   timestamptz,
  planned_end    timestamptz,
  ingested_at    timestamptz not null default now(),
  primary key (report_date, resource_name, unit_code, outage_start)
);
create index if not exists resource_outages_start_idx on resource_outages (outage_start);
alter table resource_outages enable row level security;
drop policy if exists resource_outages_read on resource_outages;
create policy resource_outages_read on resource_outages for select using (true);
grant select on resource_outages to anon, authenticated;
"""

INSERT = """
insert into resource_outages
  (report_date, resource_name, unit_code, fuel_type, outage_type,
   max_mw, available_mw, reduction_mw, outage_start, planned_end)
values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
on conflict (report_date, resource_name, unit_code, outage_start) do update set
  available_mw = excluded.available_mw,
  reduction_mw = excluded.reduction_mw,
  planned_end = excluded.planned_end
"""


def num(x):
    try:
        return float(x) if x not in (None, "", " ") else None
    except (TypeError, ValueError):
        return None


def ts(x):
    if isinstance(x, datetime):
        return x.replace(tzinfo=timezone.utc) if x.tzinfo is None else x
    return None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--days", type=int, default=1400)
    args = ap.parse_args()

    client = ErcotClient()

    def headers():
        return {"Authorization": f"Bearer {client.token()}",
                "Ocp-Apim-Subscription-Key": os.environ["ERCOT_SUBSCRIPTION_KEY"]}

    def raw(u, tries=5):
        for a in range(tries):
            try:
                req = urllib.request.Request(u, headers=headers())
                with urllib.request.urlopen(req, timeout=180) as r:
                    return r.read()
            except Exception:
                if a == tries - 1:
                    raise
                time.sleep(6 * (a + 1))

    cutoff = datetime.now(timezone.utc) - timedelta(days=args.days)
    docs, page = [], 1
    while True:
        d = json.loads(raw("https://api.ercot.com/api/public-reports/archive/"
                           f"NP1-346-ER?size=1000&page={page}"))
        arcs = d.get("archives") or []
        docs.extend(a for a in arcs
                    if datetime.fromisoformat(a["postDatetime"]).replace(tzinfo=timezone.utc) >= cutoff)
        if not arcs or datetime.fromisoformat(arcs[-1]["postDatetime"]).replace(tzinfo=timezone.utc) < cutoff:
            break
        page += 1
    print(f"daily snapshots in window: {len(docs)}", flush=True)

    with psycopg.connect(os.environ["DATABASE_URL"], connect_timeout=30) as conn:
        conn.execute("set statement_timeout='10min'")
        with conn.cursor() as cur:
            cur.execute(DDL)
        conn.commit()
        total = 0
        for i, a in enumerate(reversed(docs)):
            rd = datetime.fromisoformat(a["postDatetime"]).date()
            try:
                b = raw(a["_links"]["endpoint"]["href"])
                z = zipfile.ZipFile(io.BytesIO(b))
                wb = openpyxl.load_workbook(io.BytesIO(z.read(z.namelist()[0])),
                                            read_only=True, data_only=True)
            except Exception as exc:
                print(f"  [{i+1}/{len(docs)}] {rd} FAILED {str(exc)[:70]}", flush=True)
                continue
            sheet = next((wb[s] for s in wb.sheetnames if "Outages" in s), None)
            if sheet is None:
                continue
            rows, header_seen = [], False
            for r in sheet.iter_rows(values_only=True):
                if not header_seen:
                    if r and str(r[0]).strip() == "Resource Name":
                        header_seen = True
                    continue
                if not r or not r[0]:
                    continue
                rows.append((rd, str(r[0]).strip(), str(r[1] or "").strip(),
                             str(r[2] or "").strip() or None,
                             str(r[3] or "").strip() or None,
                             num(r[4]), num(r[5]), num(r[6]), ts(r[7]), ts(r[8])))
            if rows:
                with conn.cursor() as cur:
                    cur.executemany(INSERT, rows)
                conn.commit()
                total += len(rows)
            if (i + 1) % 50 == 0:
                print(f"  [{i+1}/{len(docs)}] rows {total:,}", flush=True)
        with conn.cursor() as cur:
            cur.execute("""select count(*), min(report_date), max(report_date),
                                  count(distinct resource_name) from resource_outages""")
            print("\nresource_outages:", cur.fetchone())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
