#!/usr/bin/env python3
"""Ingest ERCOT's GIS Report (generation interconnection queue), monthly.

    python scripts/ingest_gis_queue.py --parse-only   # smoke test, no DB
    python scripts/ingest_gis_queue.py                # latest report -> DB
    python scripts/ingest_gis_queue.py --limit 3      # newest 3 reports

The GIS Report is MIS report type 15933 (EMIL pg7-200-er, Public) on
www.ercot.com — same public listing/download servlets as the CRR results in
ercot/crr.py, but listed via the JSON servlet rather than scraping HTML. The
listing mixes GIS reports with Co-located Battery reports under the same
type id, so filter on the GIS_Report filename prefix.

Sheets parsed: "Project Details - Large Gen" (full milestone columns) and
"Project Details - Small Gen" (short form). Column positions have drifted
between vintages, so both are resolved by header name, not index. The
"1-1-1900" placeholder dates ERCOT uses for n/a become NULL.
"""
from __future__ import annotations

import argparse
import io
import pathlib
import re
import sys
from datetime import date, datetime

from dotenv import load_dotenv

ROOT = pathlib.Path(__file__).resolve().parents[1]
load_dotenv(ROOT / ".env")
sys.path.insert(0, str(ROOT))

import httpx  # noqa: E402
import openpyxl  # noqa: E402

LISTING = "https://www.ercot.com/misapp/servlets/IceDocListJsonWS"
DOWNLOAD = "https://www.ercot.com/misdownload/servlets/mirDownload"
REPORT_TYPE_ID = "15933"
TIMEOUT = httpx.Timeout(300.0, connect=30.0)

MONTHS = {m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"], 1)}

COLUMNS = ["report_month", "inr", "sheet", "project_name", "study_phase",
           "entity", "poi_location", "county", "zone", "projected_cod",
           "fuel", "technology", "capacity_mw", "screening_started",
           "screening_complete", "fis_requested", "fis_approved", "ia_signed",
           "construction_start", "construction_end", "approved_energization",
           "approved_synchronization", "comment"]

INSERT = f"""
insert into interconnection_queue ({", ".join(COLUMNS)})
values ({", ".join(["%s"] * len(COLUMNS))})
on conflict (report_month, inr) do update set
  sheet = excluded.sheet, project_name = excluded.project_name,
  study_phase = excluded.study_phase, entity = excluded.entity,
  poi_location = excluded.poi_location, county = excluded.county,
  zone = excluded.zone, projected_cod = excluded.projected_cod,
  fuel = excluded.fuel, technology = excluded.technology,
  capacity_mw = excluded.capacity_mw,
  screening_started = excluded.screening_started,
  screening_complete = excluded.screening_complete,
  fis_requested = excluded.fis_requested,
  fis_approved = excluded.fis_approved, ia_signed = excluded.ia_signed,
  construction_start = excluded.construction_start,
  construction_end = excluded.construction_end,
  approved_energization = excluded.approved_energization,
  approved_synchronization = excluded.approved_synchronization,
  comment = excluded.comment
"""

INR_RX = re.compile(r"^\d{2}INR", re.I)


def list_gis_reports() -> list[dict]:
    """GIS report documents on the MIS, newest first."""
    resp = httpx.get(LISTING, params={"reportTypeId": REPORT_TYPE_ID},
                     timeout=TIMEOUT, follow_redirects=True)
    resp.raise_for_status()
    docs = resp.json()["ListDocsByRptTypeRes"]["DocumentList"]
    out = []
    for d in docs:
        doc = d["Document"]
        if doc.get("FriendlyName", "").startswith("GIS_Report"):
            out.append(doc)
    return out


def report_month(friendly_name: str) -> date | None:
    """GIS_Report_July2026 / GIS_Report_Jun2026 -> 2026-07-01 / 2026-06-01.

    ERCOT is not consistent about abbreviating the month, so match any prefix
    of a month name of three letters or more.
    """
    m = re.search(r"GIS_Report_?([A-Za-z]+)_?(\d{4})", friendly_name)
    if not m:
        return None
    frag, year = m.group(1).lower(), int(m.group(2))
    for name, num in MONTHS.items():
        if name.startswith(frag) or frag.startswith(name[:3]):
            return date(year, num, 1)
    return None


def _text(v) -> str | None:
    if v is None:
        return None
    s = str(v).replace("_x000D_", "\n").strip()
    return s or None


def _date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        d = v.date()
    elif isinstance(v, date):
        d = v
    else:
        try:
            d = datetime.strptime(str(v).strip()[:10], "%Y-%m-%d").date()
        except ValueError:
            try:
                d = datetime.strptime(str(v).strip(), "%m/%d/%Y").date()
            except ValueError:
                return None
    return None if d.year <= 1900 else d


def _num(v) -> float | None:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return None


def parse_sheet(ws, month: date, kind: str) -> list[tuple]:
    idx: dict[str, int] = {}
    rows: list[tuple] = []
    for row in ws.iter_rows(values_only=True):
        if not idx:
            if row and _text(row[0]) == "INR":
                idx = {_text(v) or "": i for i, v in enumerate(row) if v}
            continue
        inr = _text(row[0] if row else None)
        if not inr or not INR_RX.match(inr):
            continue

        def col(name):
            i = idx.get(name)
            return row[i] if i is not None and i < len(row) else None

        rows.append((
            month, inr, kind, _text(col("Project Name")),
            _text(col("GIM Study Phase")), _text(col("Interconnecting Entity")),
            _text(col("POI Location")), _text(col("County")),
            _text(col("CDR Reporting Zone")), _date(col("Projected COD")),
            _text(col("Fuel")), _text(col("Technology")),
            _num(col("Capacity (MW)")),
            _date(col("Screening Study Started")),
            _date(col("Screening Study Complete")),
            _date(col("FIS Requested")), _date(col("FIS Approved")),
            _date(col("IA Signed")),
            _date(col("Construction Start")), _date(col("Construction End")),
            _date(col("Approved for Energization")),
            _date(col("Approved for Synchronization")),
            _text(col("Comment")),
        ))
    return rows


def parse_workbook(blob: bytes, month: date) -> list[tuple]:
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True)
    rows: list[tuple] = []
    for sheet_name, kind in (("Project Details - Large Gen", "large"),
                             ("Project Details - Small Gen", "small")):
        if sheet_name in wb.sheetnames:
            rows.extend(parse_sheet(wb[sheet_name], month, kind))
    wb.close()
    uniq = {(r[0], r[1]): r for r in rows}   # INR is the identity; last wins
    return list(uniq.values())


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=1,
                    help="how many recent monthly reports to ingest")
    ap.add_argument("--parse-only", action="store_true",
                    help="download and parse, print a summary, write nothing")
    args = ap.parse_args()

    docs = list_gis_reports()[:args.limit]
    if not docs:
        raise SystemExit("no GIS_Report documents in the MIS listing")

    all_rows: list[tuple] = []
    for doc in docs:
        month = report_month(doc["FriendlyName"])
        if month is None:
            print(f"skip {doc['FriendlyName']}: cannot read report month")
            continue
        resp = httpx.get(DOWNLOAD, params={"doclookupId": doc["DocID"]},
                         timeout=TIMEOUT, follow_redirects=True)
        resp.raise_for_status()
        rows = parse_workbook(resp.content, month)
        large = sum(1 for r in rows if r[2] == "large")
        print(f"{doc['FriendlyName']}: {len(rows)} projects "
              f"({large} large gen, {len(rows) - large} small gen), "
              f"report_month={month}")
        all_rows.extend(rows)

    for r in all_rows[:2]:
        print("sample:", dict(zip(COLUMNS, r)))

    if args.parse_only:
        return 0

    import os
    import psycopg
    with psycopg.connect(os.environ["DATABASE_URL"], connect_timeout=30) as conn:
        conn.execute("set statement_timeout='10min'")
        with conn.cursor() as cur:
            cur.executemany(INSERT, all_rows)
        conn.commit()
        with conn.cursor() as cur:
            cur.execute("select report_month, count(*) from interconnection_queue "
                        "group by 1 order by 1 desc limit 5")
            print("interconnection_queue:", cur.fetchall())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
