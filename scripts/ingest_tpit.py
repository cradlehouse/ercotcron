#!/usr/bin/env python3
"""Ingest ERCOT's TPIT report (Transmission Project Information Tracking).

    python scripts/ingest_tpit.py --parse-only     # smoke test, no DB
    python scripts/ingest_tpit.py                  # full ingest (needs table)

TPIT is ERCOT's workbook of every 60kV+ transmission project with a material
impact on power flow: in-service dates, counties, status, voltage. Public
(EMIL zp8-801-m), posted as a plain document link on
https://www.ercot.com/gridinfo/planning — no MIS report type, no stable URL.
The filename changes with each release (quarterly-ish, plus ad-hoc updates),
so ingest scrapes the page for the anchor titled "Transmission Project and
Information Tracking" and takes whatever it points to.

Each release is loaded as a full snapshot keyed by report_date + sheet:
projects migrate between the Future/Planned/Completed/Cancelled sheets over
time, and that migration is the signal. The release date is recovered from
the sheet names themselves (e.g. FutureTPIT071326NoCost -> 2026-07-13).
"""
from __future__ import annotations

import argparse
import io
import pathlib
import re
import sys
from datetime import date, datetime

from dotenv import load_dotenv

ROOT = pathlib.Path(__file__).resolve().parents[1]
load_dotenv(ROOT / ".env")
sys.path.insert(0, str(ROOT))

import httpx  # noqa: E402
import openpyxl  # noqa: E402

PLANNING_PAGE = "https://www.ercot.com/gridinfo/planning"
ANCHOR_TITLE = "Transmission Project and Information Tracking"
TIMEOUT = httpx.Timeout(180.0, connect=30.0)

# Sheet name carries both the kind and the release date: FutureTPIT071326NoCost
SHEET_RX = re.compile(r"^(Future|Planned|Completed|Cancelled)TPIT(\d{6})", re.I)

COLUMNS = ["report_date", "sheet", "project_number", "phase", "title",
           "description", "comments", "from_location", "to_location", "status",
           "owner", "owner_project", "projected_in_service",
           "actual_in_service", "voltage_kv", "miles_new", "miles_rebuilt",
           "autotransformer_mva", "reactive_mvar", "county_from", "county_to",
           "tier", "rpg_number"]

INSERT = f"""
insert into transmission_projects ({", ".join(COLUMNS)})
values ({", ".join(["%s"] * len(COLUMNS))})
on conflict (report_date, sheet, project_number, phase)
do update set title = excluded.title, description = excluded.description,
  comments = excluded.comments, from_location = excluded.from_location,
  to_location = excluded.to_location, status = excluded.status,
  owner = excluded.owner, owner_project = excluded.owner_project,
  projected_in_service = excluded.projected_in_service,
  actual_in_service = excluded.actual_in_service,
  voltage_kv = excluded.voltage_kv, miles_new = excluded.miles_new,
  miles_rebuilt = excluded.miles_rebuilt,
  autotransformer_mva = excluded.autotransformer_mva,
  reactive_mvar = excluded.reactive_mvar,
  county_from = excluded.county_from, county_to = excluded.county_to,
  tier = excluded.tier, rpg_number = excluded.rpg_number
"""


def find_workbook_url() -> str:
    resp = httpx.get(PLANNING_PAGE, timeout=TIMEOUT, follow_redirects=True)
    resp.raise_for_status()
    # The anchor carries a title attribute; match on that rather than the
    # filename, which is different every release.
    m = re.search(
        r'href="([^"]+)"[^>]*title="' + re.escape(ANCHOR_TITLE) + '"',
        resp.text,
    )
    if not m:
        raise SystemExit(f"no '{ANCHOR_TITLE}' link on {PLANNING_PAGE}")
    url = m.group(1)
    return url if url.startswith("http") else "https://www.ercot.com" + url


def _text(v) -> str | None:
    if v is None:
        return None
    # TSP contact cells embed _x000D_ carriage returns from Excel
    s = str(v).replace("_x000D_", "\n").strip()
    return s or None


def _date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        d = v.date()
    elif isinstance(v, date):
        d = v
    else:
        s = str(v).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m/%Y"):
            try:
                d = datetime.strptime(s, fmt).date()
                break
            except ValueError:
                continue
        else:
            return None
    # 1900-01-01 is ERCOT's placeholder for "not applicable"
    return None if d.year <= 1900 else d


def _num(v) -> float | None:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return None


def parse_workbook(blob: bytes) -> list[tuple]:
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True)
    rows: list[tuple] = []
    for name in wb.sheetnames:
        m = SHEET_RX.match(name)
        if not m:
            continue
        kind = m.group(1).lower()
        report_date = datetime.strptime(m.group(2), "%m%d%y").date()
        ws = wb[name]
        header_seen = False
        for row in ws.iter_rows(values_only=True):
            first = _text(row[0] if row else None)
            if not header_seen:
                header_seen = first == "ERCOT Project Number"
                continue
            if not first:
                continue
            c = list(row) + [None] * 33
            rows.append((
                report_date, kind, first, _text(c[30]) or "",
                _text(c[1]), _text(c[2]), _text(c[3]), _text(c[4]),
                _text(c[5]), _text(c[6]), _text(c[8]), _text(c[10]),
                _date(c[11]), _date(c[12]), _num(c[13]), _num(c[14]),
                _num(c[15]), _num(c[16]), _num(c[17]), _text(c[18]),
                _text(c[19]), _text(c[20]), _text(c[21]),
            ))
    wb.close()
    # Same project+phase can appear twice in a sheet (data-entry dupes);
    # last one wins, matching the upsert's behaviour.
    uniq = {(r[0], r[1], r[2], r[3]): r for r in rows}
    return list(uniq.values())


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--url", help="workbook URL (default: scrape planning page)")
    ap.add_argument("--parse-only", action="store_true",
                    help="download and parse, print a summary, write nothing")
    args = ap.parse_args()

    url = args.url or find_workbook_url()
    print("source:", url, flush=True)
    resp = httpx.get(url, timeout=TIMEOUT, follow_redirects=True)
    resp.raise_for_status()
    rows = parse_workbook(resp.content)

    by_sheet: dict[str, int] = {}
    for r in rows:
        by_sheet[r[1]] = by_sheet.get(r[1], 0) + 1
    print(f"parsed {len(rows)} projects, report_date={rows[0][0] if rows else '?'}, "
          f"by sheet: {by_sheet}")
    for r in rows[:2]:
        print("sample:", dict(zip(COLUMNS, r)))

    if args.parse_only:
        return 0

    import os
    import psycopg
    with psycopg.connect(os.environ["DATABASE_URL"], connect_timeout=30) as conn:
        conn.execute("set statement_timeout='10min'")
        with conn.cursor() as cur:
            cur.executemany(INSERT, rows)
        conn.commit()
        with conn.cursor() as cur:
            cur.execute("select count(*), max(report_date) from transmission_projects")
            print("transmission_projects:", cur.fetchone())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
